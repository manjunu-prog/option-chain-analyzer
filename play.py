import streamlit as st
import requests
import pandas as pd
import time
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
import io

# ──────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────
DHAN_ACCESS_TOKEN = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiJ9.eyJwX2lwIjoiIiwic19pcCI6IiIsImlzcyI6ImRoYW4iLCJwYXJ0bmVySWQiOiIiLCJleHAiOjE3Nzg1NTY3MjQsImlhdCI6MTc3ODQ3MDMyNCwidG9rZW5Db25zdW1lclR5cGUiOiJTRUxGIiwid2ViaG9va1VybCI6Imh0dHBzOi8vd2ViLmRoYW4uY28vaW5kZXgvcHJvZmlsZSIsImRoYW5DbGllbnRJZCI6IjExMDgwNjYwOTQifQ.IkDHPIOhASygIrmrv8g_W1awS4mLx7Vzds4-AK67omL-UmwrbuaSXWO7WHBl4ToTcPGjnhCH1_gI3_kYIUGbcQ"
DHAN_CLIENT_ID    = "1108066094"
TELEGRAM_TOKEN   = "7851529826:AAHfyHVrVZi5iQubljaNgde76gPhr8pxql4"
TELEGRAM_CHAT_ID = "567677761"

API_BASE        = "https://api.dhan.co/v2"
OPTIONCHAIN_URL = f"{API_BASE}/optionchain"
EXPIRY_LIST_URL = f"{API_BASE}/optionchain/expirylist"

UNDERLYING_MAP = {
    "NIFTY":  {"Scrip": 13, "Segments": ["IDX_I", "NSE_FNO"], "step": 50},
    "SENSEX": {"Scrip": 1,  "Segments": ["BSE_FNO", "IDX_I"], "step": 100},
}

def send_telegram_combined_analysis(index_name, ltp, atm, pcr, df, step):
    try:
        # Filter ATM ±5 and sort
        df = df[(df["STRIKE"] >= atm - step*5) & (df["STRIKE"] <= atm + step*5)]
        df = df.sort_values("STRIKE", ascending=False)

        # Increase height to accommodate dual bars (Volume + OI Change)
        width, height = 850, 80 + len(df)*60 + 40
        img = Image.new("RGB", (width, height), (10, 12, 18)) # Darker background
        draw = ImageDraw.Draw(img)
        font = ImageFont.load_default()

        # Header
        draw.text((20, 15), f"🚀 {index_name} DUAL ANALYSIS | LTP: {ltp:,.0f} | PCR: {pcr:.2f}", fill=(255,255,255), font=font)
        draw.text((20, 35), "LEFT: CALL (Vol/ΔOI) | RIGHT: PUT (Vol/ΔOI)", fill=(150, 150, 150), font=font)

        # Normalize scaling
        max_vol = max(df["_cv"].max(), df["_pv"].max(), 1)
        max_oi  = max(df["_cd"].abs().max(), df["_pd"].abs().max(), 1) # Use absolute for ΔOI scaling

        y = 75
        bar_max_w = 220

        for _, r in df.iterrows():
            strike = int(r["STRIKE"])
            # Data points
            cv, pv = r["_cv"], r["_pv"]
            cd, pd = r["_cd"], r["_pd"]

            # --- 1. VOLUME BARS (Top thin bar) ---
            cv_w = int((cv / max_vol) * bar_max_w)
            pv_w = int((pv / max_vol) * bar_max_w)
            # CE Vol (Grey)
            draw.rectangle([20, y, 20 + cv_w, y + 8], fill=(100, 116, 139))
            # PE Vol (Grey)
            draw.rectangle([width - 20 - pv_w, y, width - 20, y + 8], fill=(100, 116, 139))

            # --- 2. OI CHANGE BARS (Bottom thick bar) ---
            cd_w = int((abs(cd) / max_oi) * bar_max_w)
            pd_w = int((abs(pd) / max_oi) * bar_max_w)
            
            # CE ΔOI Bar Color: Red if positive (selling), Green if negative (covering)
            ce_color = (239, 68, 68) if cd > 0 else (34, 197, 94)
            draw.rectangle([20, y + 12, 20 + cd_w, y + 25], fill=ce_color)
            
            # PE ΔOI Bar Color: Green if positive (selling/support), Red if negative
            pe_color = (34, 197, 94) if pd > 0 else (239, 68, 68)
            draw.rectangle([width - 20 - pd_w, y + 12, width - 20, y + 25], fill=pe_color)

            # --- 3. STRIKE TEXT (Center) ---
            strike_color = (255, 255, 255)
            txt = f"{strike} ATM" if strike == atm else str(strike)
            if strike == atm:
                draw.rectangle([width//2 - 50, y, width//2 + 50, y + 25], outline=(255,255,255))
            
            draw.text((width//2 - 30, y + 5), txt, fill=strike_color, font=font)

            # --- 4. VALUES ---
            # Volume labels
            draw.text((20 + cv_w + 5, y), f"V:{cv/1e5:.1f}L", fill=(148, 163, 184), font=font)
            draw.text((width - 20 - pv_w - 70, y), f"V:{pv/1e5:.1f}L", fill=(148, 163, 184), font=font)
            # OI Delta labels
            draw.text((20 + cd_w + 5, y + 12), f"Δ:{cd/1e5:.1f}L", fill=ce_color, font=font)
            draw.text((width - 20 - pd_w - 70, y + 12), f"Δ:{pd/1e5:.1f}L", fill=pe_color, font=font)

            y += 55 # Space for next strike

        # Save and Send
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        requests.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendPhoto",
                      data={"chat_id": TELEGRAM_CHAT_ID},
                      files={"photo": ("analysis.png", buf, "image/png")}, timeout=15)
    except Exception as e:
        print(f"Error generating dual chart: {e}")



def send_telegram_strikewise_image(index_name, ltp, atm, pcr, df, step):
    try:
        # Filter ATM ±5
        df = df[(df["STRIKE"] >= atm - step*5) & (df["STRIKE"] <= atm + step*5)]
        df = df.sort_values("STRIKE", ascending=False)

        width, height = 800, 50 + len(df)*40 + 40
        img = Image.new("RGB", (width, height), (15, 18, 25))
        draw = ImageDraw.Draw(img)

        font = ImageFont.load_default()

        # Title
        draw.text((20, 10), f"{index_name} | LTP: {ltp:,.0f}", fill=(255,255,255), font=font)

        max_vol = max(df["_cv"].max(), df["_pv"].max(), 1)

        y = 50
        bar_max_width = 200

        for _, r in df.iterrows():
            strike = int(r["STRIKE"])
            c_vol = r["_cv"]
            p_vol = r["_pv"]
            c_delta = r["_cd"]
            p_delta = r["_pd"]

            # sentiment color
            if c_delta > p_delta:
                color = (255, 80, 80)   # red
            elif p_delta > c_delta:
                color = (80, 255, 120)  # green
            else:
                color = (200, 200, 200)

            # CE bar (left)
            c_width = int((c_vol / max_vol) * bar_max_width)
            draw.rectangle([20, y, 20 + c_width, y + 15], fill=(180,180,180))

            # PE bar (right)
            p_width = int((p_vol / max_vol) * bar_max_width)
            draw.rectangle([width - 20 - p_width, y, width - 20, y + 15], fill=(180,180,180))

            # strike text
            if strike == atm:
                txt = f"{strike} ATM"
            else:
                txt = f"{strike}"

            # center text
            draw.text((width//2 - 40, y), txt, fill=color, font=font)

            # values
            draw.text((20 + c_width + 5, y), f"{c_vol/1e5:.1f}L", fill=(255,255,255), font=font)
            draw.text((width - 20 - p_width - 60, y), f"{p_vol/1e5:.1f}L", fill=(255,255,255), font=font)

            y += 35

        # PCR
        draw.text((20, height - 30), f"PCR: {pcr:.2f}", fill=(255,255,255), font=font)

        # Save to bytes
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)

        # Send to Telegram
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendPhoto",
            data={"chat_id": TELEGRAM_CHAT_ID},
            files={"photo": ("oc.png", buf, "image/png")},
            timeout=10
        )

    except Exception as e:
        pass


# ──────────────────────────────────────────────
# TELEGRAM ALERT
# ──────────────────────────────────────────────
def send_telegram_alert(index_name, ltp, atm, expiry, pcr, df):
    try:
        # ── CALL side ──
        max_c_oi_chg_row = df.loc[df["_cd"].idxmax()]
        max_c_vol_row    = df.loc[df["_cv"].idxmax()]

        # ── PUT side ──
        max_p_oi_chg_row = df.loc[df["_pd"].idxmax()]
        max_p_vol_row    = df.loc[df["_pv"].idxmax()]

        c_arrow = "▲" if max_c_oi_chg_row["_cd"] >= 0 else "▼"
        p_arrow = "▲" if max_p_oi_chg_row["_pd"] >= 0 else "▼"

        msg = (
            f"📊 *{index_name} Option Chain Alert*\n"
            f"🕐 {time.strftime('%d-%b %H:%M')} | Expiry: {expiry}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"💰 LTP: `{ltp:,.0f}` | ATM: `{int(atm)}` | PCR: `{pcr:.2f}`\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"📈 *CALL (CE)*\n"
            f"  🔹 Highest OI Chg : `{int(max_c_oi_chg_row['STRIKE'])}` — ΔOI: `{max_c_oi_chg_row['_cd']/1e5:.2f}L {c_arrow}` | LTP: `{max_c_oi_chg_row['C LTP']}`\n"
            f"  🔹 Highest Vol    : `{int(max_c_vol_row['STRIKE'])}` — Vol: `{max_c_vol_row['_cv']/1e5:.2f}L` | LTP: `{max_c_vol_row['C LTP']}`\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"📉 *PUT (PE)*\n"
            f"  🔸 Highest OI Chg : `{int(max_p_oi_chg_row['STRIKE'])}` — ΔOI: `{max_p_oi_chg_row['_pd']/1e5:.2f}L {p_arrow}` | LTP: `{max_p_oi_chg_row['P LTP']}`\n"
            f"  🔸 Highest Vol    : `{int(max_p_vol_row['STRIKE'])}` — Vol: `{max_p_vol_row['_pv']/1e5:.2f}L` | LTP: `{max_p_vol_row['P LTP']}`\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"_Auto-alert on every page refresh_"
        )

        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
            json={"chat_id": TELEGRAM_CHAT_ID, "text": msg, "parse_mode": "Markdown"},
            timeout=5
        )
    except Exception as e:
        pass  # Silent fail — never break the dashboard

def send_telegram_alertMSG(index_name, ltp, atm, expiry, pcr, df):
    try:
        # ── CALL side ──
        max_c_oi_chg_row = df.loc[df["_cd"].idxmax()]
        max_c_vol_row    = df.loc[df["_cv"].idxmax()]

        # ── PUT side ──
        max_p_oi_chg_row = df.loc[df["_pd"].idxmax()]
        max_p_vol_row    = df.loc[df["_pv"].idxmax()]

        c_arrow = "▲" if max_c_oi_chg_row["_cd"] >= 0 else "▼"
        p_arrow = "▲" if max_p_oi_chg_row["_pd"] >= 0 else "▼"

        msg = (
            f"📊 *{index_name} Option Chain Alert*\n"
            f"🕐 {time.strftime('%d-%b %H:%M')} | Expiry: {expiry}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"💰 LTP: `{ltp:,.0f}` | ATM: `{int(atm)}` | PCR: `{pcr:.2f}`\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"📈 *CALL (CE)*\n"
            f"  🔹 Highest OI Chg : `{int(max_c_oi_chg_row['STRIKE'])}` — ΔOI: `{max_c_oi_chg_row['_cd']/1e5:.2f}L {c_arrow}` | LTP: `{max_c_oi_chg_row['C LTP']}`\n"
            f"  🔹 Highest Vol    : `{int(max_c_vol_row['STRIKE'])}` — Vol: `{max_c_vol_row['_cv']/1e5:.2f}L` | LTP: `{max_c_vol_row['C LTP']}`\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"📉 *PUT (PE)*\n"
            f"  🔸 Highest OI Chg : `{int(max_p_oi_chg_row['STRIKE'])}` — ΔOI: `{max_p_oi_chg_row['_pd']/1e5:.2f}L {p_arrow}` | LTP: `{max_p_oi_chg_row['P LTP']}`\n"
            f"  🔸 Highest Vol    : `{int(max_p_vol_row['STRIKE'])}` — Vol: `{max_p_vol_row['_pv']/1e5:.2f}L` | LTP: `{max_p_vol_row['P LTP']}`\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"_Auto-alert on every page refresh_"
        )

        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
            json={"chat_id": TELEGRAM_CHAT_ID, "text": msg, "parse_mode": "Markdown"},
            timeout=5
        )
    except Exception as e:
        pass  # Silent fail — never break the dashboard


# ──────────────────────────────────────────────
# EXCEL ALERT — colored xlsx sent to Telegram
# ──────────────────────────────────────────────
def send_excel_to_telegram(index_name, ltp, atm, expiry, pcr, df,
                            c_vol_top3, c_oi_top3, p_vol_top3, p_oi_top3,
                            min_c_oi_idx, min_p_oi_idx,
                            c_neg_oi_top3=None, p_neg_oi_top3=None):
    if c_neg_oi_top3 is None: c_neg_oi_top3 = []
    if p_neg_oi_top3 is None: p_neg_oi_top3 = []
    try:
        display_cols = ["C OI CH%","C VOL (L)","CALL OI (L)","C Δ OI","C LTP",
                        "STRIKE","IV","P LTP","P Δ OI","PUT OI (L)","P VOL (L)","P OI CH%"]
        export_df = df[display_cols].copy()

        wb = Workbook()
        ws = wb.active
        ws.title = f"{index_name} OC"

        # ── Color map ──
        def fill(hex_col):
            return PatternFill("solid", fgColor=hex_col.replace("#",""))

        FILLS = {
            "CYAN1":  fill("#1976d2"), "CYAN2":  fill("#64b5f6"), "CYAN3":  fill("#bbdefb"),
            "PINK1":  fill("#c62828"), "PINK2":  fill("#ef5350"), "PINK3":  fill("#ffcdd2"),
            "YELLOW": fill("#ffe082"), "YELLOW2": fill("#ffd54f"), "YELLOW3": fill("#fff9c4"), "WHITE":  fill("#ffffff"),
            "STRIKE": fill("#c8dff5"), "DARK":   fill("#f0f6ff"),
            "HEADER": fill("#daeaf8"),
        }
        WHITE_FONT  = Font(color="0D1B2A", bold=True, name="Calibri", size=10)
        BLACK_FONT  = Font(color="0D1B2A", bold=True, name="Calibri", size=10)
        NORMAL_FONT = Font(color="0D1B2A", name="Calibri", size=10)
        GREY_FONT   = Font(color="2C5F8A", name="Calibri", size=10)
        CENTER      = Alignment(horizontal="center", vertical="center")
        thin        = Side(style="thin", color="B8D4F0")
        BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

        col_idx = {col: i+1 for i, col in enumerate(display_cols)}

        # ── Title row ──
        ws.merge_cells(f"A1:{get_column_letter(len(display_cols))}1")
        title_cell = ws["A1"]
        title_cell.value = (f"{index_name}  |  LTP: {ltp:,.0f}  |  ATM: {int(atm)}  "
                            f"|  PCR: {pcr:.2f}  |  Expiry: {expiry}  "
                            f"|  {time.strftime('%d-%b-%Y %H:%M')}")
        title_cell.fill    = FILLS["HEADER"]
        title_cell.font    = Font(color="0D1B2A", bold=True, name="Calibri", size=11)
        title_cell.alignment = CENTER
        ws.row_dimensions[1].height = 22

        # ── Header row ──
        for ci, col in enumerate(display_cols, 1):
            cell = ws.cell(row=2, column=ci, value=col)
            cell.fill      = FILLS["HEADER"]
            cell.font      = Font(color="1A3A5C", bold=True, name="Calibri", size=10)
            cell.alignment = CENTER
            cell.border    = BORDER
        ws.row_dimensions[2].height = 18

        # ── Data rows ──
        for ri, (_, row) in enumerate(export_df.iterrows(), 3):
            ws.row_dimensions[ri].height = 18
            strike_val = df.loc[ri-3, "STRIKE"]
            is_atm     = (strike_val == atm)
            df_idx     = ri - 3

            for ci, col in enumerate(display_cols, 1):
                cell = ws.cell(row=ri, column=ci, value=row[col])
                cell.alignment = CENTER
                cell.border    = BORDER

                # Default dark bg
                cell.fill = FILLS["WHITE"] if is_atm else FILLS["DARK"]
                cell.font = BLACK_FONT     if is_atm else NORMAL_FONT

                # STRIKE column
                if col == "STRIKE":
                    cell.fill = FILLS["WHITE"] if is_atm else FILLS["STRIKE"]
                    cell.font = BLACK_FONT if is_atm else WHITE_FONT
                # IV column
                elif col == "IV":
                    cell.fill = FILLS["STRIKE"]
                    cell.font = GREY_FONT

                # CE Vol highlights
                elif col == "C VOL (L)":
                    if df_idx == c_vol_top3[0]:
                        cell.fill, cell.font = FILLS["CYAN1"], BLACK_FONT
                    elif len(c_vol_top3) > 1 and df_idx == c_vol_top3[1]:
                        cell.fill, cell.font = FILLS["CYAN2"], WHITE_FONT
                    elif len(c_vol_top3) > 2 and df_idx == c_vol_top3[2]:
                        cell.fill, cell.font = FILLS["CYAN3"], WHITE_FONT

                # CE OI Change highlights
                elif col == "C Δ OI":
                    if len(c_neg_oi_top3) > 2 and df_idx == c_neg_oi_top3[2]:
                        cell.fill, cell.font = FILLS["YELLOW3"], BLACK_FONT
                    if len(c_neg_oi_top3) > 1 and df_idx == c_neg_oi_top3[1]:
                        cell.fill, cell.font = FILLS["YELLOW2"], BLACK_FONT
                    if len(c_neg_oi_top3) > 0 and df_idx == c_neg_oi_top3[0]:
                        cell.fill, cell.font = FILLS["YELLOW"], BLACK_FONT
                    elif df_idx == c_oi_top3[0]:
                        cell.fill, cell.font = FILLS["CYAN1"], BLACK_FONT
                    elif len(c_oi_top3) > 1 and df_idx == c_oi_top3[1]:
                        cell.fill, cell.font = FILLS["CYAN2"], WHITE_FONT
                    elif len(c_oi_top3) > 2 and df_idx == c_oi_top3[2]:
                        cell.fill, cell.font = FILLS["CYAN3"], WHITE_FONT

                # PE OI Change highlights
                elif col == "P Δ OI":
                    if len(p_neg_oi_top3) > 2 and df_idx == p_neg_oi_top3[2]:
                        cell.fill, cell.font = FILLS["YELLOW3"], BLACK_FONT
                    if len(p_neg_oi_top3) > 1 and df_idx == p_neg_oi_top3[1]:
                        cell.fill, cell.font = FILLS["YELLOW2"], BLACK_FONT
                    if len(p_neg_oi_top3) > 0 and df_idx == p_neg_oi_top3[0]:
                        cell.fill, cell.font = FILLS["YELLOW"], BLACK_FONT
                    elif df_idx == p_oi_top3[0]:
                        cell.fill, cell.font = FILLS["PINK1"], BLACK_FONT
                    elif len(p_oi_top3) > 1 and df_idx == p_oi_top3[1]:
                        cell.fill, cell.font = FILLS["PINK2"], WHITE_FONT
                    elif len(p_oi_top3) > 2 and df_idx == p_oi_top3[2]:
                        cell.fill, cell.font = FILLS["PINK3"], WHITE_FONT

                # PE Vol highlights
                elif col == "P VOL (L)":
                    if df_idx == p_vol_top3[0]:
                        cell.fill, cell.font = FILLS["PINK1"], BLACK_FONT
                    elif len(p_vol_top3) > 1 and df_idx == p_vol_top3[1]:
                        cell.fill, cell.font = FILLS["PINK2"], WHITE_FONT
                    elif len(p_vol_top3) > 2 and df_idx == p_vol_top3[2]:
                        cell.fill, cell.font = FILLS["PINK3"], WHITE_FONT

        # ── Column widths ──
        col_widths = [9,10,12,14,8,9,6,8,14,12,10,9]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── Save to bytes and send ──
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        fname = f"{index_name}_OC_{time.strftime('%Y%m%d_%H%M')}.xlsx"
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument",
            data={"chat_id": TELEGRAM_CHAT_ID,
                  "caption": f"📊 {index_name} Option Chain | {time.strftime('%d-%b %H:%M')}"},
            files={"document": (fname, buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=15
        )
    except Exception as e:
        pass  # Silent — never break dashboard


def send_telegram_strikewise(index_name, ltp, atm, pcr, df, step):
    try:
        msg_lines = []
        msg_lines.append(f"📊 *{index_name}* | LTP: `{ltp:,.0f}`\n")

        # Filter for ATM ±5 strikes and sort
        df = df[(df["STRIKE"] >= atm - step*5) & (df["STRIKE"] <= atm + step*5)]
        df_sorted = df.sort_values("STRIKE", ascending=False)

        def short_lakh(val):
            return f"{val / 1e5:.1f}L"

        for _, r in df_sorted.iterrows():
            strike = int(r["STRIKE"])
            c_vol_raw = r["_cv"]
            p_vol_raw = r["_pv"]

            # Icon logic based on Volume instead of Delta OI
            # 🟢 Green: Call Volume > Put Volume
            # 🔴 Red: Put Volume > Call Volume
            if c_vol_raw > p_vol_raw:
                icon = "🟢"
            elif p_vol_raw > c_vol_raw:
                icon = "🔴"
            else:
                icon = "⚪"

            # Strike label formatting
            if strike == atm:
                strike_txt = f"{icon} {strike} ATM"
            else:
                strike_txt = f"{icon} {strike}"

            # Formatting values (Volume and LTP only)
            c_vol = short_lakh(c_vol_raw)
            p_vol = short_lakh(p_vol_raw)
            
            # Ensure LTP columns match your DataFrame keys (usually 'C LTP' and 'P LTP')
            c_ltp = f"{float(r['C LTP']):.0f}"
            p_ltp = f"{float(r['P LTP']):.0f}"

            # 👉 FINAL CLEAN LINE: [Call Vol/LTP] [Strike] [LTP/Put Vol]
            line = (
                f"`{c_vol}/{c_ltp:<3}`  "
                f"{strike_txt:^14}  "
                f"`{p_ltp:>3}/{p_vol}`"
            )

            msg_lines.append(line)

        msg_lines.append(f"\nPCR: `{pcr:.2f}`")
        final_msg = "\n".join(msg_lines)

        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
            json={
                "chat_id": TELEGRAM_CHAT_ID,
                "text": final_msg,
                "parse_mode": "Markdown"
            },
            timeout=10
        )

    except Exception as e:
        print(f"Error in send_telegram_strikewise: {e}")

# ──────────────────────────────────────────────
# CSS - CLEAN DARK TERMINAL
# ──────────────────────────────────────────────
st.set_page_config(layout="wide", page_title="Pro Option Terminal", page_icon="📊")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;700;800&family=JetBrains+Mono:wght@500;700&display=swap');

/* Light Blue/Red Theme Background */
[data-testid="stAppViewContainer"], [data-testid="stHeader"], .main {
    background-color: #f0f6ff !important;
}

/* Dark Text for Light Background */
html, body, [class*="css"] {
    color: #0d1b2a !important;
    font-family: 'Inter', sans-serif !important;
}

/* Table Headers */
.section-headers { display: grid; grid-template-columns: 1fr 110px 1fr; gap: 10px; margin-bottom: 5px; }
.sh { 
    text-align: center; padding: 10px; font-weight: 700; border-radius: 4px; 
    border: 1px solid #b8d4f0; background: #daeaf8; font-size: 0.8rem; color: #0d1b2a;
}

/* Dataframe Container */
[data-testid="stDataFrameResizable"] {
    background-color: #f0f6ff !important;
    border: 1px solid #b8d4f0 !important;
}
.stDataFrame th {
    background-color: #c8dff5 !important;
    color: #1a3a5c !important;
    font-size: 0.7rem !important;
}

/* Style for Buttons */
div.stButton > button {
    width: 100%;
    background-color: #daeaf8;
    color: #0d1b2a;
    border: 1px solid #7ab3e0;
}
div.stButton > button:hover {
    border-color: #1976d2;
    color: #1976d2;
}
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────
def _headers():
    return {"access-token": DHAN_ACCESS_TOKEN, "client-id": DHAN_CLIENT_ID, "Content-Type": "application/json"}

def fmt_lakh(val): return f"{val/1e5:.1f}"

# ──────────────────────────────────────────────
# REFRESH & INDEX SELECTION
# ──────────────────────────────────────────────
if "index_choice" not in st.session_state:
    st.session_state.index_choice = "NIFTY"

# 3 Minute Refresh (180 Seconds)
refresh_interval = 420
if "last_refresh" not in st.session_state: 
    st.session_state.last_refresh = time.time()

elapsed = time.time() - st.session_state.last_refresh
if elapsed >= refresh_interval:
    st.session_state.last_refresh = time.time()
    st.rerun()

# Index Selection Buttons at the very top
col_btn1, col_btn2, col_spacer = st.columns([1, 1, 5])
with col_btn1:
    if st.button("NIFTY"):
        st.session_state.index_choice = "NIFTY"
        st.rerun()
with col_btn2:
    if st.button("SENSEX"):
        st.session_state.index_choice = "SENSEX"
        st.rerun()

cfg = UNDERLYING_MAP[st.session_state.index_choice]

# ──────────────────────────────────────────────
# DATA FETCHING
# ──────────────────────────────────────────────
found_expiry, used_seg = None, None
for seg in cfg["Segments"]:
    r_exp = requests.post(EXPIRY_LIST_URL, json={"UnderlyingScrip": cfg["Scrip"], "UnderlyingSeg": seg}, headers=_headers())
    exp_list = r_exp.json().get("data", [])
    if exp_list:
        found_expiry, used_seg = exp_list[0], seg
        break

if found_expiry:
    r_oc = requests.post(OPTIONCHAIN_URL, json={"UnderlyingScrip": cfg["Scrip"], "UnderlyingSeg": used_seg, "Expiry": found_expiry}, headers=_headers())
    data_sec = r_oc.json().get("data", {})
    oc_map   = data_sec.get("oc", {})
    ltp      = float(data_sec.get("last_price") or 0)

    if oc_map:
        atm = round(ltp / cfg["step"]) * cfg["step"]
        rows = []
        for strike_s, legs in oc_map.items():
            strike_f = float(strike_s)
            if abs(strike_f - atm) <= (cfg["step"] * 10):
                ce, pe = legs.get("ce", {}), legs.get("pe", {})
                c_delta = int(ce.get("oi", 0)) - int(ce.get("previous_oi") or 0)
                p_delta = int(pe.get("oi", 0)) - int(pe.get("previous_oi") or 0)
                rows.append({
                    "C OI CH%": f"{ce.get('oi_change_pct', 0):.1f}%",
                    "C VOL (L)": f"{int(ce.get('volume',0))/1e5:.2f}",
                    "CALL OI (L)": f"{int(ce.get('oi',0))/1e5:.2f}",
                    "C Δ OI": f"{c_delta:,} {'▲' if c_delta >= 0 else '▼'}",
                    "C LTP": f"{float(ce.get('last_price', 0)):.1f}",
                    "STRIKE": strike_f,
                    "IV": f"{float(ce.get('iv', 0)):.1f}" if ce.get('iv') else "0.0",
                    "P LTP": f"{float(pe.get('last_price', 0)):.1f}",
                    "P Δ OI": f"{p_delta:,} {'▲' if p_delta >= 0 else '▼'}",
                    "PUT OI (L)": f"{int(pe.get('oi',0))/1e5:.2f}",
                    "P VOL (L)": f"{int(pe.get('volume',0))/1e5:.2f}",
                    "P OI CH%": f"{pe.get('oi_change_pct', 0):.1f}%",
                    "_cv": int(ce.get("volume", 0)), "_pv": int(pe.get("volume", 0)),
                    "_cd": c_delta, "_pd": p_delta,
                    "_coi": int(ce.get("oi", 0)), "_poi": int(pe.get("oi", 0))
                })

        df = pd.DataFrame(rows).sort_values("STRIKE").reset_index(drop=True)
        
        # PCR Calculation
        total_c_oi = df["_coi"].sum()
        total_p_oi = df["_poi"].sum()
        pcr = total_p_oi / total_c_oi if total_c_oi else 0

        # ── Highlight indices (computed once, used by both Excel and table) ──
        c_vol_top3   = df['_cv'].nlargest(3).index.tolist()
        c_oi_top3    = df['_cd'].nlargest(3).index.tolist()
        p_vol_top3   = df['_pv'].nlargest(3).index.tolist()
        p_oi_top3    = df['_pd'].nlargest(3).index.tolist()
        min_c_oi_idx = df['_cd'].idxmin()
        min_p_oi_idx = df['_pd'].idxmin()
        # Top 3 most negative OI changes (CE and PE)
        c_neg_oi_top3 = df[df['_cd'] < 0]['_cd'].nsmallest(3).index.tolist()
        p_neg_oi_top3 = df[df['_pd'] < 0]['_pd'].nsmallest(3).index.tolist()

        # ──────────────────────────────────────────────
        # TELEGRAM ALERT — fires on every page load/refresh
        # ──────────────────────────────────────────────
        #send_telegram_alert(
        #    st.session_state.index_choice, ltp, atm,
        #    found_expiry, pcr, df
        #)

        # ── Excel to Telegram ──
        send_excel_to_telegram(
            st.session_state.index_choice, ltp, atm, found_expiry, pcr, df,
            c_vol_top3, c_oi_top3, p_vol_top3, p_oi_top3,
            min_c_oi_idx, min_p_oi_idx,
            c_neg_oi_top3, p_neg_oi_top3
        )

        send_telegram_strikewise(
        st.session_state.index_choice,
        ltp,
        atm,
        pcr,
        df,
        cfg["step"]
        )

        send_telegram_strikewise_image(
        st.session_state.index_choice,
        ltp,
        atm,
        pcr,
        df,
        cfg["step"]
        )
        #Rename columns to match the new function's requirements
        df_for_telegram = df.rename(columns={
        "STRIKE": "STRIKE", 
        "CE Volume": "_cv", 
        "PE Volume": "_pv", 
        "CE Δ OI": "_cd", 
        "PE Δ OI": "_pd"
        })

        # Call the new dual-analysis function
        #send_telegram_combined_analysis(
        #index_name=st.session_state.index_choice,
        #ltp=ltp,
        #atm=atm,
        #pcr=pcr,
        #df=df_for_telegram,
        #step=cfg["step"]
        #)
        

        # ──────────────────────────────────────────────
        # TOP PANEL (METRICS)
        # ──────────────────────────────────────────────
        st.markdown(f"""
        <div style="background-color: #daeaf8; padding: 10px 0px; border-bottom: 1px solid #7ab3e0;">
            <h1 style="color: #0d1b2a; font-size: 2.2rem; font-weight: 800; margin-bottom: 20px; letter-spacing: -1px;">
                NSE {st.session_state.index_choice} | ATM {int(atm)} | LTP {ltp:,.0f} | {time.strftime('%H:%M')}
            </h1>
            <div style="display: grid; grid-template-columns: repeat(6, 1fr); gap: 15px;">
                <div><div style="color: #2c5f8a; font-size: 0.75rem; font-weight: 600;">LTP</div><div style="font-size: 1.6rem; font-weight: 700; color: #0d1b2a;">{ltp:,.0f}</div></div>
                <div><div style="color: #2c5f8a; font-size: 0.75rem; font-weight: 600;">ATM</div><div style="font-size: 1.6rem; font-weight: 700; color: #0d1b2a;">{int(atm)}</div></div>
                <div><div style="color: #2c5f8a; font-size: 0.75rem; font-weight: 600;">PCR</div><div style="font-size: 1.6rem; font-weight: 700; color: #0d1b2a;">{pcr:.2f}</div></div>
                <div><div style="color: #2c5f8a; font-size: 0.75rem; font-weight: 600;">CE OI</div><div style="font-size: 1.6rem; font-weight: 700; color: #0d1b2a;">{fmt_lakh(total_c_oi)}L</div></div>
                <div><div style="color: #2c5f8a; font-size: 0.75rem; font-weight: 600;">PE OI</div><div style="font-size: 1.6rem; font-weight: 700; color: #0d1b2a;">{fmt_lakh(total_p_oi)}L</div></div>
                <div><div style="color: #2c5f8a; font-size: 0.75rem; font-weight: 600;">CE OI Chg</div><div style="font-size: 1.6rem; font-weight: 700; color: #0d1b2a;">{fmt_lakh(df['_cd'].sum())}L</div></div>
            </div>
            <div style="color: #3a6ea5; font-size: 0.7rem; margin-top: 20px;">
                Expiry: {found_expiry} | Update in: {int(refresh_interval - elapsed)}s
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div style='margin-bottom: 20px;'></div>", unsafe_allow_html=True)
        st.markdown("<div class='section-headers'><div class='sh'>CALLS</div><div class='sh'>STRIKE</div><div class='sh'>PUTS</div></div>", unsafe_allow_html=True)

        # ──────────────────────────────────────────────
        # CSV DOWNLOAD
        # ──────────────────────────────────────────────
        csv_data = df[["C OI CH%","C VOL (L)","CALL OI (L)","C Δ OI","C LTP","STRIKE","IV","P LTP","P Δ OI","PUT OI (L)","P VOL (L)","P OI CH%"]].to_csv(index=False)
        st.download_button(
            label="⬇️ Download CSV",
            data=csv_data,
            file_name=f"{st.session_state.index_choice}_OC_{time.strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv"
        )

        # ──────────────────────────────────────────────
        # HIGHLIGHT LOGIC
        # ──────────────────────────────────────────────
        # Cyan shades: #1 bright, #2 medium, #3 dim
        CYAN1, CYAN2, CYAN3 = '#1976d2', '#64b5f6', '#bbdefb'
        PINK1, PINK2, PINK3 = '#c62828', '#ef5350', '#ffcdd2'
        YELLOW1, YELLOW2, YELLOW3 = '#ffe082', '#ffd54f', '#fff9c4'  # yellow shades for negative OI

        def style_terminal(data):
            styles = pd.DataFrame('', index=data.index, columns=data.columns)
            styles.update(pd.DataFrame('background-color: #f0f6ff; color: #0d1b2a;', index=data.index, columns=data.columns))
            styles['STRIKE'] = 'background-color: #c8dff5; color: #0d1b2a; font-weight: 700;'
            styles['IV']     = 'background-color: #daeaf8; color: #3a6ea5;'

            # ── CE Vol — 1st/2nd/3rd ──
            for rank, (idx, bg) in enumerate(zip(c_vol_top3, [CYAN1, CYAN2, CYAN3])):
                fg = '#000000' if rank == 0 else '#ffffff'
                styles.loc[idx, 'C VOL (L)'] = f'background-color: {bg}; color: {fg}; font-weight: 700;'

            # ── CE OI Change — 1st/2nd/3rd positive ──
            for rank, (idx, bg) in enumerate(zip(c_oi_top3, [CYAN1, CYAN2, CYAN3])):
                fg = '#000000' if rank == 0 else '#ffffff'
                styles.loc[idx, 'C Δ OI'] = f'background-color: {bg}; color: {fg}; font-weight: 700;'

            # ── CE top 3 NEGATIVE OI change — yellow shades ──
            for rank, (idx, bg) in enumerate(zip(c_neg_oi_top3, [YELLOW1, YELLOW2, YELLOW3])):
                styles.loc[idx, 'C Δ OI'] = f'background-color: {bg}; color: #000000; font-weight: 700;'

            # ── PE Vol — 1st/2nd/3rd ──
            for rank, (idx, bg) in enumerate(zip(p_vol_top3, [PINK1, PINK2, PINK3])):
                fg = '#000000' if rank == 0 else '#ffffff'
                styles.loc[idx, 'P VOL (L)'] = f'background-color: {bg}; color: {fg}; font-weight: 700;'

            # ── PE OI Change — 1st/2nd/3rd positive ──
            for rank, (idx, bg) in enumerate(zip(p_oi_top3, [PINK1, PINK2, PINK3])):
                fg = '#000000' if rank == 0 else '#ffffff'
                styles.loc[idx, 'P Δ OI'] = f'background-color: {bg}; color: {fg}; font-weight: 700;'

            # ── PE top 3 NEGATIVE OI change — yellow shades ──
            for rank, (idx, bg) in enumerate(zip(p_neg_oi_top3, [YELLOW1, YELLOW2, YELLOW3])):
                styles.loc[idx, 'P Δ OI'] = f'background-color: {bg}; color: #000000; font-weight: 700;'

            # ── ATM Strike ──
            atm_idx = data[data['STRIKE'] == atm].index
            if not atm_idx.empty:
                styles.loc[atm_idx[0], 'STRIKE'] = 'background-color: #ffffff; color: #000000; font-weight: 900;'
            return styles

        display_cols = ["C OI CH%", "C VOL (L)", "CALL OI (L)", "C Δ OI", "C LTP", "STRIKE", "IV", "P LTP", "P Δ OI", "PUT OI (L)", "P VOL (L)", "P OI CH%"]
        raw_cols = ["_cv", "_pv", "_cd", "_pd", "_coi", "_poi"]

        st.dataframe(
            df[display_cols + raw_cols].style.apply(style_terminal, axis=None)
            .format(precision=0).hide(axis="columns", subset=raw_cols),
            use_container_width=True, height=780
        )

# Forced reload via component to sync with refresh_interval
st.components.v1.html(f"<script>setTimeout(function(){{ window.parent.location.reload(); }}, {refresh_interval * 1000});</script>", height=0)
